#!/usr/bin/env python3
"""Generate the companies_template.xlsx example file."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Companies"

# Headers
headers = [
    "Company Name", "URL", "Category", "Niche", "Description",
    "Domain Rating", "Monthly Traffic", "Keywords", "Authority Score",
]

header_fill = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

# Example rows
examples = [
    ["Klaviyo", "https://www.klaviyo.com", "Email/SMS Automation", "Marketing automation", "Email and SMS marketing platform for ecommerce", 89, 1500000, 45000, 78],
    ["Gorgias", "https://www.gorgias.com", "Customer Engagement", "Helpdesk", "Customer service helpdesk for ecommerce", 76, 350000, 12000, 65],
    ["ShipBob", "https://www.shipbob.com", "Fulfilment", "3PL", "Third-party logistics and order fulfilment", 82, 900000, 28000, 72],
    ["Recharge", "https://rechargepayments.com", "Payments", "Subscriptions", "Subscription billing platform for ecommerce", 71, 200000, 8500, 61],
    ["Yotpo", "https://www.yotpo.com", "Marketing", "Reviews & UGC", "Reviews, loyalty, and user-generated content platform", 85, 650000, 22000, 74],
]

for row_idx, example in enumerate(examples, 2):
    for col_idx, value in enumerate(example, 1):
        ws.cell(row=row_idx, column=col_idx, value=value)

# Set column widths
widths = [20, 35, 25, 25, 50, 15, 15, 12, 15]
for col, width in enumerate(widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

wb.save("companies_template.xlsx")
print("Created companies_template.xlsx")
