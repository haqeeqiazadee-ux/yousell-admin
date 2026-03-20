"""Export per-company analysis results to a structured Excel file."""

import logging
import os
from collections import defaultdict

logger = logging.getLogger("gap_analyzer")

# Sections we aggregate across all companies for the Summary tab
_DIM_SECTIONS = [
    ("dim2_functionality_tech", "Functionality & Technology"),
    ("dim3_content_messaging", "Content & Messaging"),
    ("dim4_services_products", "Services & Products"),
    ("dim5_business_model", "Business Model"),
]

# Fields within each dimension that hold the most insightful data
_DIM_FIELDS = {
    "dim2_functionality_tech": [
        ("core_product", "Core Products"),
        ("key_features", "Key Features (across ecosystem)"),
        ("integrations", "Integration Patterns"),
        ("product_maturity", "Maturity Distribution"),
        ("gap_for_your_project", "Gaps For Your Project"),
    ],
    "dim3_content_messaging": [
        ("primary_message", "Messaging Approaches"),
        ("messaging_clarity", "Clarity Distribution"),
        ("seo_depth", "SEO Patterns"),
        ("social_proof", "Social Proof Patterns"),
        ("primary_cta", "CTA Patterns"),
        ("gap_for_your_project", "Gaps For Your Project"),
    ],
    "dim4_services_products": [
        ("pricing_model", "Pricing Models"),
        ("pricing_visibility", "Pricing Visibility"),
        ("packaging", "Packaging Approaches"),
        ("upsell_mechanics", "Upsell Mechanics"),
        ("gap_for_your_project", "Gaps For Your Project"),
    ],
    "dim5_business_model": [
        ("business_model", "Business Models"),
        ("revenue_model", "Revenue Models"),
        ("icp", "Target ICPs"),
        ("gtm_motion", "GTM Motions"),
        ("competitive_position", "Competitive Positioning"),
        ("gap_for_your_project", "Gaps For Your Project"),
    ],
}


def export_to_excel(cache_data: dict, output_path: str) -> bool:
    """Export all company analyses to an Excel workbook with one row per company."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        logger.error("[ERROR] openpyxl not installed. Run: pip install openpyxl")
        return False

    companies = cache_data.get("companies", {})
    if not companies:
        logger.warning("[WARN] No companies to export")
        return False

    wb = Workbook()

    # ─── Sheet 1: Full Analysis ───
    ws = wb.active
    ws.title = "Company Analysis"

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
    wrap = Alignment(wrap_text=True, vertical="top")
    thin_border = Border(
        left=Side(style="thin", color="D5D5D5"),
        right=Side(style="thin", color="D5D5D5"),
        top=Side(style="thin", color="D5D5D5"),
        bottom=Side(style="thin", color="D5D5D5"),
    )

    headers = [
        "Domain",
        "Company Name",
        "URL",
        "Category",
        "Niche",
        "Status",
        "Source",
        "One-Line Verdict",
        # Dim 2: Functionality
        "Core Product",
        "Key Features",
        "Tech Signals",
        "Integrations",
        "Product Maturity",
        "Functionality Gap",
        # Dim 3: Content
        "Primary Message",
        "Messaging Clarity",
        "Content Tone",
        "SEO Depth",
        "Social Proof",
        "Primary CTA",
        "Content Gap",
        # Dim 4: Services
        "Product Catalogue",
        "Pricing Model",
        "Pricing Visibility",
        "Packaging",
        "Upsell Mechanics",
        "Services Gap",
        # Dim 5: Business Model
        "Business Model",
        "Revenue Model",
        "ICP",
        "GTM Motion",
        "Competitive Position",
        "Growth Stage",
        "Business Model Gap",
        # Lists
        "Top Opportunities",
        "Value-Add Ideas",
        "Watch Out For",
        # Meta
        "Pages Scraped",
        "Analysed At",
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Freeze header row
    ws.freeze_panes = "A2"

    row_idx = 2
    for domain, data in sorted(companies.items()):
        analysis = data.get("analysis", data)
        status = data.get("status", "unknown")
        source = data.get("source", "")

        dim2 = analysis.get("dim2_functionality_tech", {}) or {}
        dim3 = analysis.get("dim3_content_messaging", {}) or {}
        dim4 = analysis.get("dim4_services_products", {}) or {}
        dim5 = analysis.get("dim5_business_model", {}) or {}

        row = [
            domain,
            analysis.get("company_name", domain),
            analysis.get("url", ""),
            analysis.get("category", ""),
            analysis.get("niche", ""),
            status,
            source,
            analysis.get("one_line_verdict", ""),
            # Dim 2
            dim2.get("core_product", ""),
            _join_list(dim2.get("key_features", [])),
            dim2.get("tech_signals", ""),
            dim2.get("integrations", ""),
            dim2.get("product_maturity", ""),
            dim2.get("gap_for_your_project", ""),
            # Dim 3
            dim3.get("primary_message", ""),
            dim3.get("messaging_clarity", ""),
            dim3.get("content_tone", ""),
            dim3.get("seo_depth", ""),
            dim3.get("social_proof", ""),
            dim3.get("primary_cta", ""),
            dim3.get("gap_for_your_project", ""),
            # Dim 4
            dim4.get("product_catalogue", ""),
            dim4.get("pricing_model", ""),
            dim4.get("pricing_visibility", ""),
            dim4.get("packaging", ""),
            dim4.get("upsell_mechanics", ""),
            dim4.get("gap_for_your_project", ""),
            # Dim 5
            dim5.get("business_model", ""),
            dim5.get("revenue_model", ""),
            dim5.get("icp", ""),
            dim5.get("gtm_motion", ""),
            dim5.get("competitive_position", ""),
            dim5.get("growth_stage", ""),
            dim5.get("gap_for_your_project", ""),
            # Lists
            _join_list(analysis.get("top_opportunities", [])),
            _join_list(analysis.get("value_add_ideas", [])),
            _join_list(analysis.get("watch_out_for", [])),
            # Meta
            ", ".join(data.get("pages_scraped", [])),
            data.get("analysed_at", ""),
        ]

        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value else "")
            cell.alignment = wrap
            cell.border = thin_border

        row_idx += 1

    # Auto-width columns (capped at 50)
    for col_idx in range(1, len(headers) + 1):
        max_len = len(str(ws.cell(row=1, column=col_idx).value or ""))
        for r in range(2, min(row_idx, 20)):  # Sample first 18 rows
            val = str(ws.cell(row=r, column=col_idx).value or "")
            max_len = max(max_len, min(len(val), 50))
        ws.column_dimensions[_col_letter(col_idx)].width = min(max_len + 4, 50)

    # ─── Sheet 2: Opportunities Matrix ───
    ws2 = wb.create_sheet("Opportunities")
    opp_headers = ["Company", "Category", "Niche", "Opportunity"]
    for col_idx, header in enumerate(opp_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws2.freeze_panes = "A2"

    opp_row = 2
    for domain, data in sorted(companies.items()):
        analysis = data.get("analysis", data)
        name = analysis.get("company_name", domain)
        cat = analysis.get("category", "")
        niche = analysis.get("niche", "")
        for opp in analysis.get("top_opportunities", []):
            ws2.cell(row=opp_row, column=1, value=name).border = thin_border
            ws2.cell(row=opp_row, column=2, value=cat).border = thin_border
            ws2.cell(row=opp_row, column=3, value=niche).border = thin_border
            ws2.cell(row=opp_row, column=4, value=opp).border = thin_border
            opp_row += 1

    for col_idx in range(1, 5):
        ws2.column_dimensions[_col_letter(col_idx)].width = 40

    # ─── Sheet 3: Gaps Summary ───
    ws3 = wb.create_sheet("Gaps")
    gap_headers = ["Company", "Category", "Dimension", "Gap For Your Project"]
    for col_idx, header in enumerate(gap_headers, 1):
        cell = ws3.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws3.freeze_panes = "A2"

    gap_row = 2
    dim_labels = [
        ("dim2_functionality_tech", "Functionality & Tech"),
        ("dim3_content_messaging", "Content & Messaging"),
        ("dim4_services_products", "Services & Products"),
        ("dim5_business_model", "Business Model"),
    ]
    for domain, data in sorted(companies.items()):
        analysis = data.get("analysis", data)
        name = analysis.get("company_name", domain)
        cat = analysis.get("category", "")
        for dim_key, dim_label in dim_labels:
            dim_data = analysis.get(dim_key, {}) or {}
            gap = dim_data.get("gap_for_your_project", "")
            if gap and gap.lower() not in ("n/a", "not determinable", "none", ""):
                ws3.cell(row=gap_row, column=1, value=name).border = thin_border
                ws3.cell(row=gap_row, column=2, value=cat).border = thin_border
                ws3.cell(row=gap_row, column=3, value=dim_label).border = thin_border
                ws3.cell(row=gap_row, column=4, value=gap).border = thin_border
                gap_row += 1

    for col_idx in range(1, 5):
        ws3.column_dimensions[_col_letter(col_idx)].width = 40

    # ─── Sheet 4: Research Summary ───
    _build_summary_sheet(wb, companies, header_font, header_fill, wrap, thin_border)

    # Reorder: Summary as second tab
    sheet_names = wb.sheetnames
    if "Research Summary" in sheet_names:
        summary_idx = sheet_names.index("Research Summary")
        wb.move_sheet("Research Summary", offset=-(summary_idx - 1))

    # Save
    os.makedirs(os.path.dirname(os.path.abspath(output_path)) or ".", exist_ok=True)
    wb.save(output_path)
    logger.info(f"[INFO] Excel report saved: {output_path} ({row_idx - 2} companies, "
                f"{opp_row - 2} opportunities, {gap_row - 2} gaps)")
    return True


def _build_summary_sheet(wb, companies: dict, header_font, header_fill, wrap, thin_border):
    """Build a summary sheet that aggregates findings by section with verdicts."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    ws = wb.create_sheet("Research Summary")

    section_font = Font(bold=True, color="FFFFFF", size=12)
    section_fill = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
    subsection_font = Font(bold=True, color="1B2A4A", size=11)
    verdict_font = Font(bold=True, italic=True, color="2E86C1", size=10)
    normal_wrap = Alignment(wrap_text=True, vertical="top")

    col_widths = {1: 25, 2: 80, 3: 15}
    for c, w in col_widths.items():
        ws.column_dimensions[_col_letter(c)].width = w

    row = 1

    # ─── Header ───
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    cell = ws.cell(row=row, column=1, value="COMPETITIVE INTELLIGENCE — RESEARCH SUMMARY")
    cell.font = Font(bold=True, color="FFFFFF", size=14)
    cell.fill = section_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 1

    total = len(companies)
    success = sum(1 for d in companies.values() if d.get("status") == "success")
    categories = set()
    for d in companies.values():
        cat = d.get("analysis", d).get("category", "")
        if cat:
            categories.add(cat)

    ws.cell(row=row, column=1, value="Companies Analysed").font = Font(bold=True)
    ws.cell(row=row, column=2, value=f"{total} ({success} successful, across {len(categories)} categories)")
    row += 2

    # ─── Section: Top Opportunities (aggregated) ───
    row = _write_section_header(ws, row, "TOP OPPORTUNITIES ACROSS ALL COMPANIES", section_font, section_fill)
    opp_counts = defaultdict(list)  # opportunity text -> list of company names
    for domain, data in companies.items():
        analysis = data.get("analysis", data)
        name = analysis.get("company_name", domain)
        for opp in analysis.get("top_opportunities", []):
            opp_normalised = opp.strip().rstrip(".")
            opp_counts[opp_normalised].append(name)

    # Sort by frequency (most common patterns first)
    sorted_opps = sorted(opp_counts.items(), key=lambda x: -len(x[1]))
    ws.cell(row=row, column=1, value="Opportunity").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Seen At").font = Font(bold=True)
    ws.cell(row=row, column=3, value="Frequency").font = Font(bold=True)
    row += 1
    for opp_text, seen_at in sorted_opps[:50]:  # Top 50 by frequency
        ws.cell(row=row, column=1, value=opp_text).alignment = normal_wrap
        ws.cell(row=row, column=2, value=", ".join(seen_at[:10])).alignment = normal_wrap
        ws.cell(row=row, column=3, value=len(seen_at))
        row += 1
    row += 1

    # ─── Section: Value-Add Ideas (aggregated) ───
    row = _write_section_header(ws, row, "VALUE-ADD IDEAS ACROSS ALL COMPANIES", section_font, section_fill)
    idea_counts = defaultdict(list)
    for domain, data in companies.items():
        analysis = data.get("analysis", data)
        name = analysis.get("company_name", domain)
        for idea in analysis.get("value_add_ideas", []):
            idea_counts[idea.strip().rstrip(".")].append(name)

    sorted_ideas = sorted(idea_counts.items(), key=lambda x: -len(x[1]))
    ws.cell(row=row, column=1, value="Idea").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Seen At").font = Font(bold=True)
    ws.cell(row=row, column=3, value="Frequency").font = Font(bold=True)
    row += 1
    for idea_text, seen_at in sorted_ideas[:50]:
        ws.cell(row=row, column=1, value=idea_text).alignment = normal_wrap
        ws.cell(row=row, column=2, value=", ".join(seen_at[:10])).alignment = normal_wrap
        ws.cell(row=row, column=3, value=len(seen_at))
        row += 1
    row += 1

    # ─── Section: Watch-Outs (aggregated) ───
    row = _write_section_header(ws, row, "RISKS & WATCH-OUTS", section_font, section_fill)
    watch_counts = defaultdict(list)
    for domain, data in companies.items():
        analysis = data.get("analysis", data)
        name = analysis.get("company_name", domain)
        for w in analysis.get("watch_out_for", []):
            watch_counts[w.strip().rstrip(".")].append(name)

    sorted_watches = sorted(watch_counts.items(), key=lambda x: -len(x[1]))
    ws.cell(row=row, column=1, value="Risk / Watch-Out").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Seen At").font = Font(bold=True)
    ws.cell(row=row, column=3, value="Frequency").font = Font(bold=True)
    row += 1
    for w_text, seen_at in sorted_watches[:50]:
        ws.cell(row=row, column=1, value=w_text).alignment = normal_wrap
        ws.cell(row=row, column=2, value=", ".join(seen_at[:10])).alignment = normal_wrap
        ws.cell(row=row, column=3, value=len(seen_at))
        row += 1
    row += 1

    # ─── Per-Dimension Sections ───
    for dim_key, dim_title in _DIM_SECTIONS:
        row = _write_section_header(ws, row, dim_title.upper(), section_font, section_fill)

        fields = _DIM_FIELDS.get(dim_key, [])
        for field_key, field_label in fields:
            ws.cell(row=row, column=1, value=field_label).font = subsection_font
            row += 1

            # Collect all values for this field across companies
            values_by_cat = defaultdict(list)  # category -> list of (company, value)
            for domain, data in companies.items():
                analysis = data.get("analysis", data)
                dim_data = analysis.get(dim_key, {}) or {}
                val = dim_data.get(field_key, "")
                cat = analysis.get("category", "Uncategorised")
                name = analysis.get("company_name", domain)

                if isinstance(val, list):
                    for item in val:
                        item_str = str(item).strip()
                        if item_str and item_str.lower() not in ("n/a", "not determinable", "none"):
                            values_by_cat[cat].append((name, item_str))
                elif isinstance(val, str):
                    val = val.strip()
                    if val and val.lower() not in ("n/a", "not determinable", "none", ""):
                        values_by_cat[cat].append((name, val))

            if not values_by_cat:
                ws.cell(row=row, column=1, value="No data collected").alignment = normal_wrap
                row += 1
            else:
                # For short-value fields (maturity, clarity, position, visibility, revenue_model)
                # show a distribution count instead of listing every value
                is_label_field = field_key in (
                    "product_maturity", "messaging_clarity", "pricing_visibility",
                    "competitive_position", "growth_stage", "revenue_model",
                )
                if is_label_field:
                    label_counts = defaultdict(int)
                    for cat_entries in values_by_cat.values():
                        for _, v in cat_entries:
                            label_counts[v] += 1
                    dist_parts = [f"{label}: {count}" for label, count in
                                  sorted(label_counts.items(), key=lambda x: -x[1])]
                    ws.cell(row=row, column=1, value="Distribution")
                    ws.cell(row=row, column=2, value=" | ".join(dist_parts)).alignment = normal_wrap
                    row += 1
                else:
                    # For gap/text fields — group by category, show top findings
                    for cat in sorted(values_by_cat.keys()):
                        entries = values_by_cat[cat]
                        if len(entries) > 5:
                            # Summarise: show count + top examples
                            ws.cell(row=row, column=1, value=f"{cat} ({len(entries)} findings)")
                            examples = [f"{name}: {v}" for name, v in entries[:5]]
                            ws.cell(row=row, column=2,
                                    value="\n".join(examples)).alignment = normal_wrap
                            row += 1
                        else:
                            for name, v in entries:
                                ws.cell(row=row, column=1, value=f"{cat} — {name}")
                                ws.cell(row=row, column=2, value=v).alignment = normal_wrap
                                row += 1

            # Verdict row for this field
            total_findings = sum(len(v) for v in values_by_cat.values())
            if total_findings > 0 and field_key == "gap_for_your_project":
                ws.cell(row=row, column=1, value="VERDICT").font = verdict_font
                ws.cell(row=row, column=2,
                        value=f"{total_findings} gaps identified across {len(values_by_cat)} categories. "
                              f"Review the Company Analysis tab filtered by this dimension to prioritise.").font = verdict_font
                ws.cell(row=row, column=2).alignment = normal_wrap
                row += 1

            row += 1  # Spacer between fields

        row += 1  # Spacer between dimensions

    # ─── Category Breakdown ───
    row = _write_section_header(ws, row, "CATEGORY BREAKDOWN", section_font, section_fill)
    by_cat = defaultdict(list)
    for domain, data in companies.items():
        analysis = data.get("analysis", data)
        cat = analysis.get("category", "Uncategorised") or "Uncategorised"
        by_cat[cat].append(analysis)

    ws.cell(row=row, column=1, value="Category").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Summary").font = Font(bold=True)
    ws.cell(row=row, column=3, value="Count").font = Font(bold=True)
    row += 1

    for cat in sorted(by_cat.keys()):
        entries = by_cat[cat]
        all_opps = []
        for e in entries:
            all_opps.extend(e.get("top_opportunities", []))
        top_opp = all_opps[0] if all_opps else "No opportunities identified"
        verdicts = [e.get("one_line_verdict", "") for e in entries if e.get("one_line_verdict")]

        summary_parts = [f"Top opportunity: {top_opp}"]
        if verdicts:
            summary_parts.append(f"Sample verdict: {verdicts[0]}")

        ws.cell(row=row, column=1, value=cat)
        ws.cell(row=row, column=2, value=" | ".join(summary_parts)).alignment = normal_wrap
        ws.cell(row=row, column=3, value=len(entries))
        row += 1

    ws.freeze_panes = "A2"


def _write_section_header(ws, row, title, font, fill):
    """Write a merged section header row and return the next row."""
    from openpyxl.styles import Alignment
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = font
    cell.fill = fill
    cell.alignment = Alignment(horizontal="left", vertical="center")
    return row + 1


def _join_list(items: list, sep: str = "\n• ") -> str:
    """Join a list into a bullet-separated string."""
    if not items:
        return ""
    return "• " + sep.join(str(i) for i in items)


def _col_letter(col_idx: int) -> str:
    """Convert 1-based column index to Excel column letter."""
    result = ""
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result = chr(65 + remainder) + result
    return result
